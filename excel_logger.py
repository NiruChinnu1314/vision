import os
from datetime import datetime
import pytz
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,Alignment


EXCEL_PATH = "Vision.xlsx"

def append_result(vin_number: str,
                  model_name: str,
                  class_counts: dict,
                  image_path: str):

    # Current IST timestamp
    ist = pytz.timezone("Asia/Kolkata")
    current_time = datetime.now(ist).strftime("%Y-%m-%d %H:%M:%S")

    # Determine poke_yoke_status
    fixed_count = class_counts.get("fixed_bolt", 0)
    poke_yoke_status = "OK" if fixed_count == 4 else "NOT OK"

    # Create workbook if it doesn't exist
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append([
            "S.No", "VIN number", "Model",
            "Loose Bolts", "Fixed Bolts", "No Bolts",
            "poke_yoke_status", "Timestamp", "Image"
        ])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        wb.save(EXCEL_PATH)

    # Load existing workbook
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # Next serial number (header is first row)
    next_serial = ws.max_row  # since row 1 is header

    # Append new row (leave Image cell empty for now)
    ws.append([
        next_serial,
        vin_number,
        model_name,
        class_counts.get("loose_bolt", 0),
        fixed_count,
        class_counts.get("no_bolt", 0),
        poke_yoke_status,
        current_time,
        ""  # placeholder for the image
    ])
    new_row_idx = ws.max_row  # the row we just added
    for col in range(1, ws.max_column + 1):
        ws.cell(row=new_row_idx, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # Embed the detected image in the last row, last column (I)
    if os.path.exists(image_path):
        img = XLImage(image_path)
        img.width, img.height = 120, 120  # thumbnail size
        cell_ref = f"I{ws.max_row}"        # 'Image' column is I
        ws.add_image(img, cell_ref)

        # Adjust row height for the image
        ws.row_dimensions[ws.max_row].height = 100

    # Auto-adjust column widths for all columns except 'Image'
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter == 'I':  # Image column
            ws.column_dimensions[col_letter].width = 35  # wider for images
        else:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_length + 2, 12)  # minimum width 12

    wb.save(EXCEL_PATH)
    wb.close()
